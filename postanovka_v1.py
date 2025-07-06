import sys
import glob
import os
import pandas as pd
import time
import warnings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def read_file():
    try:
        files = glob.glob('*.xlsx')
        if len(files) != 1:
            raise ValueError(f"Найдено {len(files)} файлов. Нужен ровно один XLSX файл.")
        
        input_file = files[0]
        parent_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
        time_suffix = "ночь_v1" if int(time.strftime("%H")) > 12 else "день_v1"
        output_file = os.path.join(parent_dir, f"postanovka_{time.strftime('%d%m%y')}_{time_suffix}.xlsx")

        # Стили оформления
        header_font = Font(bold=True, size=12)
        category_font = Font(bold=True, size=14)
        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        df = pd.read_excel(input_file, sheet_name=0)
        df.iloc[:, 4] = df.iloc[:, 4].astype(str).str.strip()

        #Заполнение первого столбика сплошь номерами
        last_seen_values = [None, None, None, None]
        for index, row in df.iterrows():
            for i in range(4):
                if pd.notna(row.iloc[i]):
                    last_seen_values[i] = row.iloc[i]
                else:
                    row.iloc[i] = last_seen_values[i]
            df.iloc[index] = row 

        #Проверяем котопсов
        processed_combinations = set()
        previous_animal = None

        if os.path.exists('otchet.txt'):
            os.remove('otchet.txt')

        for index, row in df.iterrows():
            number = row.iloc[0]
            animal = row.iloc[3]
            condition = row.iloc[4]
            combination = (number, animal, condition)

            with open('otchet.txt', 'a', encoding='utf-8') as file:
                if combination not in processed_combinations:
                    if animal == 'Кошка':
                        if condition in ['ПЦР-РеспБС', 'ПЦР-Диар/С', 'ПЦР-ВЫБ-Соб','ПЦР-ООБСоб']:
                            warning_message = f"Котопес! Номер {int(number)} '{animal}' и '{condition}'."
                            processed_combinations.add(combination)  
                            file.write(warning_message + '\n')

                    elif animal == 'Собака':
                        if condition in ['ПЦР-РБКош', 'ПЦР-Диар/К', 'ПЦР-ВЫБ-Кош', 'ПЦР-СтомПр/К']:
                            warning_message = f"Котопес! Номер {int(number)} '{animal}' и '{condition}'."
                            processed_combinations.add(combination)
                            file.write(warning_message + '\n')

                    if (('Гемобартонеллез кошек' in condition) and (previous_animal == 'Собака')):
                        warning_message = f"Котопес! Номер {int(number)} '{previous_animal}' и '{condition}'."
                        processed_combinations.add(combination)
                        file.write(warning_message + '\n')

                    if (('Гемобартонеллез собак' in condition) and (previous_animal == 'Кошка')):
                        warning_message = f"Котопес! Номер {int(number)} '{previous_animal}' и '{condition}'."
                        processed_combinations.add(combination)
                        file.write(warning_message + '\n')

            if not pd.isna(animal):
                previous_animal = animal
        #Конец проверки котопсов

        # Сбор и обработка данных
        infections = ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью', 'Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'gibsoni', 'Микросп','Трихоф',
                      'Asperg', 'Орн', 'Полио', 'Цирко','Bruc', 'haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 
                      'Babesia spp', 'РНК', 'HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'Bord', 'Chlamyd', 
                      'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto', 'Микроспор', 'Трихоф']
        
        klesch_complex = {'ПЦР-КлещИнв': ['Ana', 'Borrel', 'Ehr', 'Babesia spp'], 'ПЦР-КлещИ':['Ana', 'Borrel', 'Ehr', 'Babesia canis']}

        results = {}
        for infection in infections:
            filtered = df[df.iloc[:, 4].str.contains(infection, na=False)]
            results[infection] = filtered.iloc[:, 0].tolist()

        #Экстрагируем нераскрытые шапки с КИ
        kleschi_numbers = set ()
        for complex_name in klesch_complex:
            complex_data = df[df.iloc[:,4].str.contains(complex_name, na=False)]
            kleschi_numbers.update(complex_data.iloc[:,0].unique().tolist())

        for num in kleschi_numbers:
                if (num not in results['Ana']):
                    results['Ana'].append(num)
                if (num not in results['Borrel']):
                    results['Borrel'].append(num)
                if (num not in results['Ehr']):
                    results['Ehr'].append(num)
                if (num not in results['Babesia spp']):
                    results['Babesia spp'].append(num)
                if num in  results['Babesia canis']:
                    results['Babesia canis'].remove(num)
                    with open('otchet.txt', 'a', encoding='utf-8') as file:
                        file.write(f'Номер {num} - cтарая шапка на B.canis!' + '\n')


        #FIV+FeLV
        fiv = set(results.get('FIV', []))
        felv = set(results.get('FeLV', []))
        results['FIV+FeLV'] = sorted(fiv & felv)
        results['FIV'] = sorted(fiv - felv)
        results['FeLV'] = sorted(felv - fiv)
        if 'РНК' in results:
            results['РНК FeLV'] = results.pop('РНК')

        formatted = {}
        for key, vals in results.items():
            chunks = [vals[i:i+8] + ['']*(8-len(vals[i:i+8])) for i in range(0, len(vals), 8)]
            for i, chunk in enumerate(chunks, 1):
                formatted[f"{key}_{i}"] = chunk
                
        results_df = pd.DataFrame(formatted)

        categories = {
            'Заяц РНК': ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью'],
            'Заяц ДНК': ['Giard', 'Salm', 'Tritri', 'Bart', 'Tox','Микросп','Трихоф', 'gibsoni', 'Asperg', 'Орн', 'Полио', 'Цирко'],
            'Genlab': ['HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'FIV+FeLV'],
            'Fractal': ['Bord', 'Chlamyd', 'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto'],
            'VectBest': ['Bruc','haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 'Babesia spp', 'РНК FeLV']
        }


        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"
        current_row = 1

        for cat_name, infections in categories.items():
            cols = [c for c in results_df.columns if any(inf in c for inf in infections)]
            if not cols:
                continue

            if current_row > 1:
                current_row += 1

            #Заголовок категории (толстые границы)
            ws.merge_cells(
                start_row=current_row,
                end_row=current_row,
                start_column=1,
                end_column=len(cols)
            )
            cell = ws.cell(row=current_row, column=1, value=cat_name)
            cell.font = category_font
            cell.alignment = Alignment(horizontal='center')
            
            for col in range(1, len(cols)+1):
                ws.cell(row=current_row, column=col).border = thick_border
            
            current_row += 1

            #Заголовки столбцов (тонкие границы)
            for col_idx, col_name in enumerate(results_df[cols].columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.border = thin_border  # Добавлено
            
            current_row += 1

            #Данные (тонкие границы)
            for _, row in results_df[cols].iterrows():
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border  # Добавлено
                current_row += 1

        #Исправленный блок автонастройки ширины
        for col in ws.columns:
            max_length = 0
            column_number = col[0].column
            for cell in col:
                if cell.row == 1:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            col_letter = get_column_letter(column_number)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_file)
        print(f"Результаты записаны в файл: {output_file}")

    except Exception as e:
        print(f"Error!{e}")
        sys.exit(1)

if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    read_file()