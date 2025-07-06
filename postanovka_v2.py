import sys
import glob
import os
import pandas as pd
import time
import warnings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Конфигурационные параметры
CATEGORIES = {
    'Заяц РНК': ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью'],
    'Заяц ДНК': ['Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'Микросп', 'Трихоф',
                'gibsoni', 'Asperg', 'Орн', 'Полио', 'Цирко'],
    'Genlab': ['HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV',
              'Camp', 'Clostr', 'FIV+FeLV'],
    'Fractal': ['Bord', 'Chlamyd', 'Crypto', 'СAV', 'M. canis', 'Ana',
               'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto'],
    'VectBest': ['Bru','haemofelis', 'haemocanis', 'perfringens', 'galiseptica',
                'Past', 'Babesia canis', 'Babesia spp', 'РНК FeLV']
}

PRIORITY_COMPLEXES = {
    'ПЦР-РБКош': ['FCV', 'HV', 'Mycoplasma spp', 'M. felis','Bord', 'Chlamyd'],
    'ПЦР-РеспБС': ['Парагрипп','HV', 'Mycoplasma spp','Bord', 'M. canis', 'СAV'],
    'ПЦР-Диар/К': ['FCoV', 'Giard', 'Salm','PV','Camp', 'Clostr','Crypto'],
    'ПЦР-Диар/С': ['CCоV','Giard', 'Salm','PV','Camp', 'Clostr', 'Crypto']
}

NON-PRIORITY-COMPLEXES = ['ПЦР-ВЫБ-Соб', 'ПЦР-ВЫБ-Кош', 'ПЦР-ООБСоб', 'ПЦР-СтомПр/К', 'ПЦР-ДОП-Список', 'ПЦР-КлещИнв', 'ПЦР-ПтицДОМЕ']

def fill_missing_values(df):
    #Заполнение пропущенных значений в первых четырех столбцах
    last_values = [None] * 4
    for idx, row in df.iterrows():
        for i in range(4):
            if pd.notna(row[i]):
                last_values[i] = row[i]
            else:
                df.iat[idx, i] = last_values[i]
    return df

def check_cat_dog_errors(df):
    #Проверка на котопсов и запись в otchet.txt"""
    errors = []
    prev_animal = None
    
    for _, row in df.iterrows():
        num = row[0]
        animal = row[3]
        condition = str(row[4])
        
        if animal == 'Кошка':
            if any(x in condition for x in ['ПЦР-РеспБС', 'ПЦР-Диар/С', 
                                            'ПЦР-ВЫБ-Соб','ПЦР-ООБСоб']):
                errors.append(f"Котопес! Номер {int(num)} '{animal}' и '{condition}'")
        
        elif animal == 'Собака':
            if any(x in condition for x in ['ПЦР-РБКош', 'ПЦР-Диар/К',
                                           'ПЦР-ВЫБ-Кош', 'ПЦР-СтомПр/К']):
                errors.append(f"Котопес! Номер {int(num)} '{animal}' и '{condition}'")
        

        if 'Гемобартонеллез кошек' in condition and prev_animal == 'Собака':
            errors.append(f"Котопес! Номер {int(num)} '{prev_animal}' и '{condition}'")
        
        if 'Гемобартонеллез собак' in condition and prev_animal == 'Кошка':
            errors.append(f"Котопес! Номер {int(num)} '{prev_animal}' и '{condition}'")
        
        if pd.notna(animal):
            prev_animal = animal
    
    if errors:
        with open('otchet.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(errors))
    
    return df


def process_priority_data(df):
    ##Обработка данных с разделением на приоритетные и обычные
    #Сбор уникальных приоритетных номеров
    priority_numbers = set()
    for complex_name in PRIORITY_COMPLEXES:
        complex_data = df[df.iloc[:,4].str.contains(complex_name, na=False)]
        priority_numbers.update(complex_data.iloc[:,0].unique().tolist())
    
    #Приоритетные инфекции (комплексы) для приоритетных номеров
    priority_results = {}
    filtered_priority = df[df.iloc[:,0].isin(priority_numbers)]
    
    for complex_name, infections in PRIORITY_COMPLEXES.items():
        for infection in infections:
            infection_data = filtered_priority[
                filtered_priority.iloc[:,4].str.contains(infection, na=False)
            ]
            numbers = infection_data.iloc[:,0].unique().tolist()
            if numbers:
                priority_results.setdefault(infection, []).extend(numbers)
    
    #Неприоритетные данные
    non_priority_df = df[~df.iloc[:,0].isin(priority_numbers)]
    other_results = {}
    all_infections = ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью', 'Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'gibsoni', 'Микросп','Трихоф',
                      'Asperg', 'Орн', 'Полио', 'Цирко', 'Bru', 'haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 
                      'Babesia spp', 'РНК', 'HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'Bord', 'Chlamyd', 
                      'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto', 'Микроспор', 'Трихоф']
    
    #Неприоритетные инфекции (одиночные) для приоритетных номеров
    priority_non_infec = {}

    for infection in all_infections:
        if infection not in [inf for lst in PRIORITY_COMPLEXES.values() for inf in lst]:
            data = filtered_priority[
                filtered_priority.iloc[:,4].str.contains(infection, na=False)
            ]
            numbers = data.iloc[:,0].unique().tolist()
            if numbers:
                priority_non_infec[infection] = numbers

            fiv = set(other_results.get('FIV', []))
            felv = set(other_results.get('FeLV', []))
            rna = set(other_results.get('РНК', []))
            other_results['FIV+FeLV'] = sorted(fiv & felv)
            other_results['FIV'] = sorted(fiv - felv)
            other_results['FeLV'] = sorted(felv - fiv)
            if rna:
                other_results['РНК FeLV'] = sorted(rna)


    #Данные для неприоритетных номеров
    for infection in all_infections:
        data = non_priority_df[
            non_priority_df.iloc[:,4].str.contains(infection, na=False)
        ]
        numbers = data.iloc[:,0].unique().tolist()
        if numbers:
            other_results[infection] = numbers
        
    fiv = set(other_results.get('FIV', []))
    felv = set(other_results.get('FeLV', []))
    rna = set(other_results.get('РНК', []))
    other_results['FIV+FeLV'] = sorted(fiv & felv)
    other_results['FIV'] = sorted(fiv - felv)
    other_results['FeLV'] = sorted(felv - fiv)
    if rna:
        other_results['РНК FeLV'] = sorted(rna)

    #print(other_results)
    
    #Объединение результатов
    return {
        'priority': {k: sorted(list(set(v))) for k, v in priority_results.items()},
        'common': {
            **{k: sorted(list(set(v))) for k, v in priority_non_infec.items()},
            **{k: sorted(list(set(v))) for k, v in other_results.items()}
        }
    }

def create_excel_report(data):
    #Создание отчета с разделением на приоритетные и обычные блоки
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    # Стили офддормления
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'),
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    category_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    
    current_row = 1
    
    def write_block(category, infections, block_type, data_dict):
        #Запись блока данных для категории
        nonlocal current_row
        cols = []
        for inf in infections:
            chunks = 1
            while f"{inf}_{chunks}" in data_dict:
                cols.append(f"{inf}_{chunks}")
                chunks += 1
        
        if not cols:
            return
        
        #Заголовок блока
        ws.merge_cells(
            start_row=current_row,
            end_row=current_row,
            start_column=1,
            end_column=len(cols)
        )
        cell = ws.cell(current_row, 1, f"{category} ({block_type})")
        cell.font = category_font
        cell.alignment = Alignment(horizontal='center')
        
        #Границы заголовка
        for col in range(1, len(cols)+1):
            ws.cell(current_row, col).border = thick_border
        
        current_row += 1
        
        #Заголовки столбцов
        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(current_row, col_idx, col.split('_')[0])
            cell.font = header_font
            cell.border = thin_border
        
        current_row += 1
        
        #Данные
        max_rows = max(len(data_dict[col]) for col in cols)
        for i in range(max_rows):
            for col_idx, col in enumerate(cols, 1):
                val = data_dict[col][i] if i < len(data_dict[col]) else ''
                ws.cell(current_row, col_idx, val).border = thin_border
            current_row += 1
        
        current_row += 1
    
    #Форматирование данных
    def format_data(data):
        formatted = {}
        for infection, numbers in data.items():
            chunks = [numbers[i:i+8] for i in range(0, len(numbers), 8)]
            for i, chunk in enumerate(chunks, 1):
                formatted[f"{infection}_{i}"] = chunk + ['']*(8-len(chunk))
        return formatted
    
    priority_data = format_data(data['priority'])
    common_data = format_data(data['common'])
    
    # Запись категорий
    for category, infections in CATEGORIES.items():
        # Приоритетные данные
        write_block(category, infections, 'Комплексы', priority_data)
        
        # Обычные данные
        write_block(category, infections, 'Не-комплексы', common_data)
    
    #Автонастройка ширины
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        if max_len > 0:
            ws.column_dimensions[get_column_letter(col[0].column)].width = (max_len + 2) * 1.2
    
    return wb

def main():
    try:
        input_files = glob.glob('*.xlsx')
        if len(input_files) != 1:
            raise ValueError(f"Найдено {len(input_files)} файлов. Требуется 1.")
        
        input_path = input_files[0]
        output_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
        time_suffix = "ночь" if int(time.strftime("%H")) > 12 else "день"
        output_path = os.path.join(
            output_dir,
            f"postanovka_{time.strftime('%d%m%y')}_{time_suffix}.xlsx"
        )
        
        df = pd.read_excel(input_path, sheet_name=0)
        df.iloc[:,4] = df.iloc[:,4].astype(str).str.strip()
        
        df = fill_missing_values(df)
        df = check_cat_dog_errors(df)
        processed_data = process_priority_data(df)
        
        report = create_excel_report(processed_data)
        report.save(output_path)
        
        print(f"Результаты записаны в файл: {output_path}")
    
    except Exception as e:
        print(f"Error!{e}")
        sys.exit(1)

if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    main()