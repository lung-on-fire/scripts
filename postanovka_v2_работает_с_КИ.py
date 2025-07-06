import sys
import glob
import os
import pandas as pd
import time
import warnings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Конфигурационные параметры
CATEGORIES = {
    'Заяц РНК': ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью'],
    'Заяц ДНК': ['Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'Микросп', 'Трихоф',
                'gibsoni', 'Asperg', 'Орн', 'Полио', 'Цирко'],
    'Genlab': ['HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV',
              'Camp', 'Clostr', 'FIV+FeLV'],
    'Fractal': ['Bord', 'Chlamyd', 'Crypto', 'СAV', 'M. canis', 'Ana',
               'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto'],
    'VectBest': ['Bruc','haemofelis', 'haemocanis', 'perfringens', 'galiseptica',
                'Past', 'Babesia canis', 'Babesia spp', 'РНК FeLV']
}

PRIORITY_COMPLEXES = {
    'ПЦР-РБКош': ['FCV', 'HV', 'Mycoplasma spp', 'M. felis','Bord', 'Chlamyd'],
    'ПЦР-РеспБС': ['Парагрипп','HV', 'Mycoplasma spp','Bord', 'M. canis', 'СAV'],
    'ПЦР-Диар/К': ['FCoV', 'Giard', 'Salm','PV','Camp', 'Clostr','Crypto'],
    'ПЦР-Диар/С': ['CCоV','Giard', 'Salm','PV','Camp', 'Clostr', 'Crypto']
}


def fill_missing_values(df):
    #Заполнение пропущенных значений в первых четырех столбцах
    last_values = [None] * 4
    for idx, row in df.iterrows():
        for i in range(4):
            if pd.notna(row.iloc[i]):
                last_values[i] = row.iloc[i]
            else:
                df.iat[idx, i] = last_values[i]
    return df

def check_cat_dog_errors(df):
    #Проверка на котопсов и запись в otchet.txt"""
    errors = []
    prev_animal = None
    if os.path.exists('otchet.txt'):
            os.remove('otchet.txt')

    for _, row in df.iterrows():
        num = row[0]
        animal = row[3]
        condition = str(row[4])
        
        if animal == 'Кошка':
            if any(x in condition for x in ['ПЦР-РеспБС', 'ПЦР-Диар/С', 
                                            'ПЦР-ВЫБ-Соб','ПЦР-ООБСоб']):
                errors.append(f"Котопес! Номер {int(num)} '{animal}' и '{condition}'")
        
        elif animal == 'Собака':
            if any(x in condition for x in ['ПЦР-РБКош', 'ПЦР-Диар/К',
                                           'ПЦР-ВЫБ-Кош', 'ПЦР-СтомПр/К']):
                errors.append(f"Котопес! Номер {int(num)} '{animal}' и '{condition}'")
        

        if 'Гемобартонеллез кошек' in condition and prev_animal == 'Собака':
            errors.append(f"Котопес! Номер {int(num)} '{prev_animal}' и '{condition}'")
        
        if 'Гемобартонеллез собак' in condition and prev_animal == 'Кошка':
            errors.append(f"Котопес! Номер {int(num)} '{prev_animal}' и '{condition}'")
        
        if pd.notna(animal):
            prev_animal = animal
    
    if errors:
        with open('otchet.txt', 'w', encoding='utf-8') as f:
            f.write('\n'.join(errors) + '\n')
    
    return df


def process_data(df):
    ##Обработка всех данных и только вывод сначала инфекций из комплексов, а потом обычных        
    infections = ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью', 'Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'gibsoni', 'Микросп','Трихоф',
                      'Asperg', 'Орн', 'Полио', 'Цирко', 'Bruc', 'haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 
                      'Babesia spp', 'РНК', 'HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'Bord', 'Chlamyd', 
                      'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto', 'Микроспор', 'Трихоф']
        
    klesch_complex = {'ПЦР-КлещИнв': ['Ana', 'Borrel', 'Ehr', 'Babesia spp'], 'ПЦР-КлещИ':['Ana', 'Borrel', 'Ehr', 'Babesia canis']}

    results = {}
    for infection in infections:
        filtered = df[df.iloc[:, 4].str.contains(infection, na=False)]
        results[infection] = filtered.iloc[:, 0].tolist()
    
    
    #Экстрагируем нераскрытые шапки с КИ
    kleschi_numbers = set ()
    for complex_name in klesch_complex:
            complex_data = df[df.iloc[:,4].str.contains(complex_name, na=False)]
            kleschi_numbers.update(complex_data.iloc[:,0].unique().tolist())

    for num in kleschi_numbers:
                if (num not in results['Ana']):
                    results['Ana'].append(num)
                if (num not in results['Borrel']):
                    results['Borrel'].append(num)
                if (num not in results['Ehr']):
                    results['Ehr'].append(num)
                if (num not in results['Babesia spp']):
                    results['Babesia spp'].append(num)
                if num in  results['Babesia canis']:
                    results['Babesia canis'].remove(num)
                    with open('otchet.txt', 'a', encoding='utf-8') as file:
                        file.write(f'Номер {num} - cтарая шапка на B.canis!' + '\n')
                                   

    #FIV+FeLV
    fiv = set(results.get('FIV', []))
    felv = set(results.get('FeLV', []))
    results['FIV+FeLV'] = sorted(fiv & felv)
    results['FIV'] = sorted(fiv - felv)
    results['FeLV'] = sorted(felv - fiv)
    if 'РНК' in results:
        results['РНК FeLV'] = results.pop('РНК')

    for complex_name, inf_list in PRIORITY_COMPLEXES.items():
        filtered_complexes = df[df.iloc[:, 4].str.contains(complex_name, na=False)]
        results[complex_name] = filtered_complexes.iloc[:, 0].tolist()

    return results

def create_excel_report(data):
    #Создание отчета с разделением на приоритетные и обычные блоки
    wb = Workbook()
    ws = wb.active
    
    # Стили офддормления
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'),
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    category_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    
    current_row = 1
    

    ####
    def write_block(category, infections, block_type, data_dict):
        #Запись блока данных для категории
        nonlocal current_row
        cols = []
        for inf in infections:
            chunks = 1
            while f"{inf}_{chunks}" in data_dict:
                cols.append(f"{inf}_{chunks}")
                chunks += 1
        
        if not cols:
            return
    ######
        
        #Заголовок блока
        ws.merge_cells(
            start_row=current_row,
            end_row=current_row,
            start_column=1,
            end_column=len(cols)
        )
        cell = ws.cell(current_row, 1, f"{category} ({block_type})")
        cell.font = category_font
        cell.alignment = Alignment(horizontal='center')
        #Границы заголовка
        for col in range(1, len(cols)+1):
            ws.cell(current_row, col).border = thick_border
        
        current_row += 1
        
        #Заголовки столбцов
        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(current_row, col_idx, col.split('_')[0])
            cell.font = header_font
            cell.border = thin_border
        
        current_row += 1
        
        #Данные
        max_rows = max(len(data_dict[col]) for col in cols)
        for i in range(max_rows):
            for col_idx, col in enumerate(cols, 1):
                val = data_dict[col][i] if i < len(data_dict[col]) else ''
                ws.cell(current_row, col_idx, val).border = thin_border
            current_row += 1
        
        current_row += 1
    
    #Форматирование данных
    def format_data(data):
        formatted = {}
        for infection, numbers in data.items():
            chunks = [numbers[i:i+8] for i in range(0, len(numbers), 8)]
            for i, chunk in enumerate(chunks, 1):
                formatted[f"{infection}_{i}"] = chunk + ['']*(8-len(chunk))
        return formatted
    

    #print(f"DATA BEFORE: {data}")
     # Собираем все приоритетные инфекции из PRIORITY_COMPLEXES
    complexes_dict = {}
    for complex_name, inf_list in PRIORITY_COMPLEXES.items():
        if complex_name in data:
            complex_numbers = data[complex_name]
            #print(complex_name)
            for inf in inf_list:
                if inf not in complexes_dict:
                    complexes_dict[inf] = []
                complexes_dict[inf].extend(complex_numbers)
            del data[complex_name]

    #print(f"DATA AFTER: {data}") ##OK
    #print(complexes_dict)


    other_data_set = set()
    complexes_set = {(key, int(value)) for key in complexes_dict for value in complexes_dict[key]}
    all_data_set = {(key, int(value)) for key in data for value in data[key]}
    other_data_set = all_data_set - complexes_set
    #print(complexes_set)

    ##
    def func_set_to_dict(cur_set):
        out_data = {}
        for key, value in cur_set:
            if key not in out_data:
                out_data[key] = []
            out_data[key].append(value)

        for key in out_data:
            out_data[key].sort()
        
        return out_data
    ###

    complexes_data = func_set_to_dict(complexes_set)
    other_data = func_set_to_dict(other_data_set)
    #print(complexes_data)
    #print(other_data)

    formatted_complex_data = format_data(complexes_data)
    formatted_other_data = format_data(other_data)

    # Обрабатываем каждую категорию
    for category, infections in CATEGORIES.items():
        #Данные комплексов
        write_block(category, infections, 'Комплексы', formatted_complex_data)

        #Данные остальные
        write_block(category, infections, 'Не-комплексы', formatted_other_data)
    
    #Автонастройка ширины
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        if max_len > 0:
            ws.column_dimensions[get_column_letter(col[0].column)].width = (max_len + 2) * 1.2
    
    return wb

def main():
    try:
        input_files = glob.glob('*.xlsx')
        if len(input_files) != 1:
            raise ValueError(f"Найдено {len(input_files)} файлов. Требуется 1.")
        
        input_path = input_files[0]
        output_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
        time_suffix = "ночь" if int(time.strftime("%H")) > 12 else "день"
        output_path = os.path.join(
            output_dir,
            f"postanovka_{time.strftime('%d%m%y')}_{time_suffix}.xlsx"
        )
        
        df = pd.read_excel(input_path, sheet_name=0)
        df.iloc[:,4] = df.iloc[:,4].astype(str).str.strip()
        
        df = fill_missing_values(df)
        df = check_cat_dog_errors(df)
        processed_data = process_data(df)
        
        report = create_excel_report(processed_data)
        report.save(output_path)
        
        print(f"Результаты записаны в файл: {output_path}")
    
    except Exception as e:
        print(f"Error!{e}")
        sys.exit(1)

if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    main()