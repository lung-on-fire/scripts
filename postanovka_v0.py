import sys, glob, os
import pandas as pd
import time
import warnings
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

def main():
    try:
        files = glob.glob('*.xlsx')
        if len(files) != 1:
            raise ValueError(f"Найдено {len(files)} XLSX файлов. Требуется ровно один файл.")
        filename = files[0]

        #Называет выходной файл в зависимости от времени системы
        if int(time.strftime("%H")) > 12:
            output_file = f"../postanovka_{time.strftime('%d%m%y')}_ночь_v0.xlsx"
        else:
            output_file = f"../postanovka_{time.strftime('%d%m%y')}_день_v0.xlsx"

        #Здесь можно указать номер листа (по умолчанию первый = 0)
        data = pd.read_excel(filename, sheet_name=0)
        data.iloc[:, 4] = data.iloc[:, 4].astype(str).str.strip()

        #Заполнение первого столбика сплошь номерами
        last_seen_values = [None, None, None, None]
        for index, row in data.iterrows():
            for i in range(4):
                if pd.notna(row.iloc[i]):
                    last_seen_values[i] = row.iloc[i]
                else:
                    row.iloc[i] = last_seen_values[i]
            data.iloc[index] = row 

        #Получаем список инфекций
        infections = ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью', 'Giard', 'Salm', 'Tritri', 'Bart', 'Tox', 'gibsoni', 'Микросп','Трихоф',
                      'Asperg', 'Орн', 'Полио', 'Цирко', 'Bruc', 'haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 
                      'Babesia spp', 'РНК', 'HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'Bord', 'Chlamyd', 
                      'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto', 'Микроспор', 'Трихоф']
        
        klesch_complex = {'ПЦР-КлещИнв': ['Ana', 'Borrel', 'Ehr', 'Babesia spp'], 'ПЦР-КлещИ':['Ana', 'Borrel', 'Ehr', 'Babesia canis']}
        
        #Проверяем котопсов
        processed_combinations = set()
        previous_animal = None

        if os.path.exists('otchet.txt'):
            os.remove('otchet.txt')

        for index, row in data.iterrows():
            number = row.iloc[0]
            animal = row.iloc[3]
            condition = row.iloc[4]
            combination = (number, animal, condition)

            with open('otchet.txt', 'a', encoding='utf-8') as file:
                if combination not in processed_combinations:
                    if animal == 'Кошка':
                        if condition in ['ПЦР-РеспБС', 'ПЦР-Диар/С', 'ПЦР-ВЫБ-Соб','ПЦР-ООБСоб']:
                            warning_message = f"Котопес! Номер {int(number)} '{animal}' и '{condition}'."
                            processed_combinations.add(combination)  
                            file.write(warning_message + '\n')

                    elif animal == 'Собака':
                        if condition in ['ПЦР-РБКош', 'ПЦР-Диар/К', 'ПЦР-ВЫБ-Кош', 'ПЦР-СтомПр/К']:
                            warning_message = f"Котопес! Номер {int(number)} '{animal}' и '{condition}'."
                            processed_combinations.add(combination)
                            file.write(warning_message + '\n')

                    if (('Гемобартонеллез кошек' in condition) and (previous_animal == 'Собака')):
                        warning_message = f"Котопес! Номер {int(number)} '{previous_animal}' и '{condition}'."
                        processed_combinations.add(combination)
                        file.write(warning_message + '\n')

                    if (('Гемобартонеллез собак' in condition) and (previous_animal == 'Кошка')):
                        warning_message = f"Котопес! Номер {int(number)} '{previous_animal}' и '{condition}'."
                        processed_combinations.add(combination)
                        file.write(warning_message + '\n')

            if not pd.isna(animal):
                previous_animal = animal
        #Конец проверки котопсов

        #Получаем номера инфекций
        results = {}

        for infection in infections:
            infection_data = data[data.iloc[:, 4].str.contains(infection, na=False)]
            infection_numbers = infection_data.iloc[:, 0]
            results[infection] = infection_numbers.tolist()


        #Экстрагируем нераскрытые шапки с КИ
        kleschi_numbers = set ()
        for complex_name in klesch_complex:
            complex_data = data[data.iloc[:,4].str.contains(complex_name, na=False)]
            kleschi_numbers.update(complex_data.iloc[:,0].unique().tolist())

        for num in kleschi_numbers:
                if (num not in results['Ana']):
                    results['Ana'].append(num)
                if (num not in results['Borrel']):
                    results['Borrel'].append(num)
                if (num not in results['Ehr']):
                    results['Ehr'].append(num)
                if (num not in results['Babesia spp']):
                    results['Babesia spp'].append(num)
                if num in  results['Babesia canis']:
                    results['Babesia canis'].remove(num) #Удаляем канис как неправильный
                    with open('otchet.txt', 'a', encoding='utf-8') as file:
                        file.write(f'Номер {num} - cтарая шапка на B.canis!' + '\n')

        #Блок кринжа чтобы разобраться с мультиплексом FIV + FeLV
        FIV_FeLV_numbers = set()
        FIV_numbers = set()
        FeLV_numbers = set()
        FeLV_RNA = set()

        for infection, numbers in results.items():
            if 'FIV' in results:
                FIV_only = [inf for inf in numbers if inf in results['FIV']]
                for number in FIV_only:
                    FIV_numbers.add(number)

            if 'FeLV' in results:
                FeLV_only = [inf for inf in numbers if inf in results['FeLV']]
                for number in FeLV_only:
                    FeLV_numbers.add(number)

            if ('FIV' and 'FeLV') in results:
                common_numbers = [inf for inf in numbers if inf in results['FIV'] and inf in results['FeLV']]
                for number in common_numbers:
                    FIV_FeLV_numbers.add(number)

            if 'РНК' in results:
                FeLV_RNA_num = [inf for inf in numbers if inf in results['РНК']]
                for number in FeLV_RNA_num:
                    FeLV_RNA.add(number)

        #Преобразуем множества в списки, удаляем старые FIV, FeLV и перезаписываем
        if (FIV_numbers | FeLV_numbers | FIV_FeLV_numbers):
            FIV_diff = FIV_numbers.difference(FIV_FeLV_numbers)
            FeLV_diff = FeLV_numbers.difference(FIV_FeLV_numbers)
            FIV_FeLV_numbers = sorted(list(FIV_FeLV_numbers))
            FIV_numbers = sorted(list(FIV_diff))
            FeLV_numbers = sorted(list(FeLV_diff))
            FeLV_RNA = sorted(list(FeLV_RNA))

            results['FIV+FeLV'] = FIV_FeLV_numbers
            results['FIV'] = FIV_numbers
            results['FeLV'] = FeLV_numbers
            if 'РНК' in results:
                results.pop('РНК')
                results['РНК FeLV'] = FeLV_RNA
        #Конец блока кринжа с мультиплексом

        #Группировка инфекций по категориям
        category_mapping = {
            'Заяц РНК': ['FCV', 'Парагрипп', 'FCoV', 'CCоV', 'Чума', 'Нью'],
            'Заяц ДНК': ['Giard', 'Salm', 'Tritri', 'Bart', 'Tox','Микросп','Трихоф', 'gibsoni', 'Asperg', 'Орн', 'Полио', 'Цирко'],
            'Genlab': ['HV', 'Mycoplasma spp', 'M. felis', 'FIV', 'FeLV', 'PV', 'Camp', 'Clostr', 'FIV+FeLV'],
            'Fractal': ['Bord', 'Chlamyd', 'Crypto', 'СAV', 'M. canis', 'Ana', 'Borrel', 'Ehr', 'Urea', 'Диро', 'Lepto'],
            'VectBest': ['Bruc','haemofelis', 'haemocanis', 'perfringens', 'galiseptica', 'Past', 'Babesia canis', 'Babesia spp', 'РНК FeLV']
        }

        #Создаем новый DataFrame для вывода
        output_data = []

        for category, infections_in_category in category_mapping.items():
            output_data.append([category] + [''] * 10) 
            for infection in infections_in_category:
                if infection in results:
                    numbers = results[infection]
                    output_data.append(['', infection] + numbers + [''] * (10 - len(numbers)))

        #Создаем Excel файл
        wb = Workbook()
        ws = wb.active

        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        thick_border = Border(
            left=Side(border_style="thick"),
            right=Side(border_style="thick"),
            top=Side(border_style="thick"),
            bottom=Side(border_style="thick")
        )

        #Записываем данные в Excel
        row_index = 1
        max_columns = max(len(row) for row in output_data)

        #Добавляем нумерацию и делаем нумерацию жирной
        column_numbers = ['', ''] + [str(i) for i in range(1, max_columns - 1)]
        ws.append(column_numbers)
        for col in range(1, max_columns + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.border = thick_border
        row_index += 1

        #Записываем остальные данные
        for row in output_data:
            ws.append(row)
            if row[0]:  #Если это строка с категорией
                #Определяем количество строк для текущей категории
                category = row[0]
                infections_in_category = category_mapping.get(category, [])
                num_rows = len(infections_in_category)
                
                #Объединяем ячейки для заголовка категории вертикально
                ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + num_rows, end_column=1)
                
                #Центрируем текст в объединенной ячейке
                cell = ws.cell(row=row_index, column=1)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)  # Жирный шрифт для заголовка категории
                
                #Применяем толстые границы для первого столбца (после объединения)
                for r in range(row_index, row_index + num_rows + 1):
                    cell = ws.cell(row=r, column=1)
                    cell.border = thick_border
            
            #Применяем жирный шрифт для подкатегорий (второй столбец)
            if row[1]:  #Если второй столбец не пустой (подкатегория)
                cell = ws.cell(row=row_index, column=2)
                cell.font = Font(bold=True)
            
            row_index += 1

        #Применяем стили границ
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        #Сохраняем файл
        wb.save(output_file)
        print(f"Результаты записаны в файл: {output_file}")

    except Exception as error:
        print(f"Error!{error}")
        sys.exit(1)

if __name__ == '__main__':
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    main()